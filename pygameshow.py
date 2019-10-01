#reference: https://www.youtube.com/watch?v=i6xMBig-pP4
#https://stackoverflow.com/questions/46390231/how-to-create-a-text-input-box-with-pygame
#have to declare pygame.init() at start for pygame to work.
#to read xlsx https://openpyxl.readthedocs.io/en/stable/
#https://stackoverflow.com/questions/34754077/openpyxl-read-only-one-column-from-excel-file-in-python
#https://www.pygame.org/wiki/TextWrap
import pygame
from openpyxl import load_workbook

questions = []
teamkey_list = []

wb = load_workbook(filename = 'questions.xlsx')

sheetname = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(sheetname)

for row in range(1,worksheet.max_row+1):
    cell_name ="A{}".format(row)
    questions.append(worksheet[cell_name].value)

pygame.init()

#make a game window
#to set full screen
#win = pygame.display.set_mode((1280,720),pygame.FULLSCREEN)
win = pygame.display.set_mode((1280,720))
winsur = pygame.display.get_surface()
#player number, can make module to edit this in future for any no of players.
player_no = 2
allplayer = {}

def Inputtextbox():
    clock = pygame.time.Clock()
    textsay = "Please enter your team name."
    textSurface = font.render(textsay, True, white)
    input_box = pygame.Rect(400, 350, 300, 34)
    color_inactive = pygame.Color('lightskyblue3')
    color_active = pygame.Color('dodgerblue2')
    color = color_inactive
    active = False
    text = ''
    done = False
    textinput = ''
    samename = ''
    samenameint = 0

    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if input_box.collidepoint(event.pos):
                    # Toggle the active variable.
                    active = not active
                else:
                    active = False
                # Change the current color of the input box.
                color = color_active if active else color_inactive
            if event.type == pygame.KEYDOWN:
                if active:
                    if event.key == pygame.K_RETURN:
                        for player in allplayer:
                            if allplayer[player][0] == text:
                                samenameint += 1
                                samename = samenameint
                        textinput = text + str(samename)
                        text = ''
                        samename = False
                        return textinput
                    elif event.key == pygame.K_BACKSPACE:
                        text = text[:-1]
                    else:
                        text += event.unicode

        win.fill(red)
        # Render the current text.
        txt_surface = font.render(text, True, color)
        # Resize the box if the text is too long.
        width = max(200, txt_surface.get_width()+10)
        input_box.w = width
        # Blit the text.
        win.blit(textSurface,(350,300))
        win.blit(txt_surface, (input_box.x+5, input_box.y+5))
        # Blit the input_box rect.
        pygame.draw.rect(win, color, input_box, 2)

        pygame.display.flip()
        clock.tick(30)

    
    
def drawText(surface, text, color, rect, font, aa=False, bkg=None):
    rect = pygame.Rect(rect)
    y = rect.top
    lineSpacing = -2

    # get the height of the font
    fontHeight = font.size("Tg")[1]

    while text:
        i = 1

        # determine if the row of text will be outside our area
        if y + fontHeight > rect.bottom:
            break

        # determine maximum width of line
        while font.size(text[:i])[0] < rect.width and i < len(text):
            i += 1

        # if we've wrapped the text, then adjust the wrap to the last word      
        if i < len(text): 
            i = text.rfind(" ", 0, i) + 1

        # render the line and blit it to the surface
        if bkg:
            image = font.render(text[:i], 1, color, bkg)
            image.set_colorkey(bkg)
        else:
            image = font.render(text[:i], aa, color)

        surface.blit(image, (rect.left, y))
        y += fontHeight + lineSpacing

        # remove the text we just blitted
        text = text[i:]

    return text

#change background color
pygame.mixer.music.load("intro.mp3")
pygame.mixer.music.play(-1,0.0)
red = [219,124,124]
white = [255,255,255]
green = [102,204,0]
grey = [160,160,160]
yellow = [255,255,0]
win.fill(red)
text = "Welcome to Bianfusia Game Show!"
text2 = "press any key to continue..."
font = pygame.font.Font('freesansbold.ttf', 30)
spacing = 40
width = 1280
height = 720
#recttop = pygame.Rect(0,0,1280,30)
#winsur.fill(white,recttop)
textSurface = font.render(text, True, white)
textSurface2 = font.render(text2, True, white)
win.blit(textSurface,(350,300))
win.blit(textSurface2,(400,450))
pygame.display.flip()

player = 1
points = 0
drawing = 40
spacing = 40
prespacing = 1240
clock = pygame.time.Clock()

while player <= player_no:
    while drawing < width:
        pygame.draw.circle(winsur, grey, (drawing,40), 20)
        pygame.draw.circle(winsur, grey, (drawing,720-40), 20)
        drawing += 80
    if spacing >= 1280:
        spacing = 40
    pygame.draw.circle(winsur, grey, (prespacing,40), 20)
    pygame.draw.circle(winsur, yellow, (spacing,40), 20)
    pygame.draw.circle(winsur, grey, (prespacing,720-40), 20)
    pygame.draw.circle(winsur, yellow, (spacing,720-40), 20)
    spacing += 80
    prespacing += 80
    if prespacing >= 1280:
        prespacing = 40
    for event in pygame.event.get():
            if event.type == pygame.KEYDOWN or event.type == pygame.MOUSEBUTTONDOWN:
                teamname = Inputtextbox()
                win.fill(red)
                text = "Welcome " + teamname + "!"
                text2 = "Please register your team key!"
                textSurface = font.render(text, True, white)
                textSurface2 = font.render(text2, True, white)
                win.blit(textSurface,(350,300))
                win.blit(textSurface2,(400,450))
                pygame.display.flip()
                done = False
                while not done:
                    for event in pygame.event.get():
                        if event.type == pygame.KEYDOWN and event.type != pygame.MOUSEBUTTONDOWN:
                            if event.key in teamkey_list and event.type != pygame.MOUSEBUTTONDOWN:
                                text = "team key taken. Please try another key."
                                textSurface = font.render(text, True, white)
                                win.blit(textSurface,(400,650))
                                pygame.display.flip()
                                continue
                            teamkey = event.key
                            teamkey_list.append(teamkey)
                            win.fill(red)
                            text = teamname + " key registered!"
                            textSurface = font.render(text, True, white)
                            win.blit(textSurface,(350,300))
                            pygame.display.flip()
                            allplayer[player] = [teamname,teamkey,points]
                            done = True
                player += 1
    pygame.display.flip()
    clock.tick(10)

x = 1
textrect = (100,100,1000,550)

while x <= len(questions):
#    questioning = False
#    while not questioning:
    team = 0
    done = False
    ans = False
    for event in pygame.event.get():
        if event.type == pygame.KEYDOWN or event.type == pygame.MOUSEBUTTONDOWN:
            pygame.mixer.pause()
            pygame.mixer.music.load("ding.mp3")
            win.fill(red)
            text = "Question " + str(x) + ": " + questions[x-1]
            drawText(winsur, text, white, textrect, font)
            pygame.display.flip()
            #questioning = True
            while not done:
                for event in pygame.event.get():
                    if event.type == pygame.KEYDOWN:
                        if event.key == allplayer[1][1]:
                            pygame.mixer.music.play(0,0.0)
                            team = 1
                            win.fill(red)
                            text = allplayer[1][0] + "!!!"
                            textSurface = font.render(text, True, white)
                            win.blit(textSurface,(350,300))
                            pygame.display.flip()
                            done = True
                        elif event.key == allplayer[2][1]:
                            pygame.mixer.music.play(0,0.0)
                            team = 2
                            win.fill(red)
                            text = allplayer[2][0] + "!!!"
                            textSurface = font.render(text, True, white)
                            win.blit(textSurface,(350,300))
                            pygame.display.flip()
                            done = True
            while not ans:
                for event in pygame.event.get():
                    if event.type == pygame.MOUSEBUTTONDOWN:
                        if event.button == 1:
                            win.fill(green)
                            text = "you are CORRECT!"
                            textSurface = font.render(text, True, white)
                            win.blit(textSurface,(350,300))
                            pygame.display.flip()
                            allplayer[team][2] += 1
                            ans = True
                            x += 1
                        elif event.button == 3:                                
                            win.fill(red)
                            text = "you are WRONG!"
                            textSurface = font.render(text, True, white)
                            win.blit(textSurface,(350,300))
                            pygame.display.flip()
                            ans = True
                            x += 1



player = 1
points = 0
drawing = 40
spacing = 40
prespacing = 1240
clock = pygame.time.Clock()
pygame.mixer.music.load("intro.mp3")
pygame.mixer.music.play(-1,0.0)

while drawing < width:
    pygame.draw.circle(winsur, grey, (drawing,40), 20)
    pygame.draw.circle(winsur, grey, (drawing,720-40), 20)
    drawing += 80

while True:
    if spacing >= 1280:
        spacing = 40
    pygame.draw.circle(winsur, grey, (prespacing,40), 20)
    pygame.draw.circle(winsur, yellow, (spacing,40), 20)
    pygame.draw.circle(winsur, grey, (prespacing,720-40), 20)
    pygame.draw.circle(winsur, yellow, (spacing,720-40), 20)
    spacing += 80
    prespacing += 80
    if prespacing >= 1280:
        prespacing = 40
    for event in pygame.event.get():
        if event.type == pygame.KEYDOWN or event.type == pygame.MOUSEBUTTONDOWN:
            win.fill(red)
            text = allplayer[1][0] + ": " + str(allplayer[1][2])
            text2 = allplayer[2][0] + ": " + str(allplayer[2][2])
            font = pygame.font.Font('freesansbold.ttf', 30)
            textSurface = font.render(text, True, white)
            textSurface2 = font.render(text2, True, white)
            win.blit(textSurface,(350,300))
            win.blit(textSurface2,(400,450))
            pygame.display.flip()
    pygame.display.flip()
    clock.tick(10)

#pygame window title.
pygame.display.set_caption("Pygame Show by Bianfusia")

#TODO:
#set full screen
#detect number and register it as buttons for game
#ask for name in pygame


